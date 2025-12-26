# BackEnd/app/main.py

import json
import uuid
import logging
from typing import List, Dict, Optional, Set
from io import BytesIO

from fastapi import FastAPI, Response, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
import openpyxl
import openpyxl.styles

# Importaciones locales
import app.seguridad as seguridad
from app.database import (
    engine, get_db, crear_tablas,
    ProfesorDB, MateriaDB, CursoDB, AulaDB, RequisitoDB, AsignacionDB, UsuarioDB,
    ConfiguracionDB
)

# --- Configuración de Logs ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Configuración Inicial ---
crear_tablas() # Esto creará 'horarios.db' según tu código
app = FastAPI()

# ==========================================
# 1. CORS
# ==========================================
origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/login")

# --- CONSTANTES ---
HORARIOS_ORDENADOS = [
  "07:40", "08:20", "09:00", "09:40", "10:20", "11:00", "11:40", "12:20",
  "13:00", "13:40", "14:20", "15:00", "15:40", "16:20", "17:00",
  "17:40", "18:20", "19:00", "19:40", "20:20", "21:00", "21:40", "22:20"
]
DIAS_SEMANA = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']

# ==========================================
# 2. MODELOS PYDANTIC
# ==========================================

class UserCreate(BaseModel):
    username: str
    password: str
    role: str = "admin" # El frontend manda 'role', nosotros lo guardaremos en 'rol'

class UsuarioRegistro(BaseModel):
    username: str
    password: str

class PasswordChange(BaseModel):
    current_password: str  # <--- CAMBIAMOS A 'current_password'
    new_password: str

class Usuario(BaseModel):
    username: str

class RequisitoCreate(BaseModel):
    curso_id: str
    materia_id: str
    profesor_id: str
    aula_preferida_id: Optional[str] = None
    horas_semanales: int

# Busca esta clase y déjala así:
class Profesor(BaseModel):
    id: Optional[str] = None
    nombre: str
    dni: Optional[str] = None  # <--- FALTABA ESTA LÍNEA
    disponibilidad: List[str]
    color: Optional[str] = "#0d9488"

class Aula(BaseModel):
    id: Optional[str] = None
    nombre: str
    tipo: Optional[str] = "Normal"
    capacidad: int = 30

class Materia(BaseModel):
    id: Optional[str] = None
    nombre: str
    color_hex: Optional[str] = "#0d9488"

class Curso(BaseModel):
    id: Optional[str] = None
    anio: str
    division: str
    cantidad_alumnos: int = 30
    nombre_display: Optional[str] = None

class Preferencias(BaseModel):
     almuerzo_slots: List[str] = []

class BusquedaSuplente(BaseModel):
    dia: str
    hora_inicio: str

class ReporteCargaHoraria(BaseModel):
    nombre_profesor: str
    horas_asignadas: int
    color: Optional[str] = "#0d9488" # <--- NUEVO

class MovimientoHorario(BaseModel):
    asignacion_id: str
    nuevo_dia: str
    nueva_hora: str

# ==========================================
# 3. SEGURIDAD Y DEPENDENCIAS
# ==========================================

def get_current_user(token: str = Depends(oauth2_scheme)) -> dict:
    payload = seguridad.verificar_token(token)
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token inválido o expirado",
            headers={"WWW-Authenticate": "Bearer"},
        )
    return {
        "username": payload.get("sub"),
        # Tu DB usa 'rol', pero el token puede tener 'role'. Normalizamos.
        "rol": payload.get("role", "admin"), 
        "force_change_password": payload.get("force_change_password", False)
    }

def get_current_admin_user(current_user: dict = Depends(get_current_user)):
    # Verificamos 'rol' (tu nombre de columna)
    if current_user["rol"] != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Requiere rol admin")
    return current_user

def calcular_hora_rango(hora_inicio: str) -> Optional[str]:
    try:
        minutos = int(hora_inicio[3:])
        hora = int(hora_inicio[:2])
        hora_fin_min = (minutos + 40) % 60
        hora_fin_hr = hora + (minutos + 40) // 60
        return f"{hora_inicio} a {hora_fin_hr:02d}:{hora_fin_min:02d}"
    except Exception:
        return None

def estan_horarios_publicados(db: Session) -> bool:
    config = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "horarios_publicados").first()
    if not config: return False
    return config.value_json == "true"


# ==========================================
# 4. ENDPOINTS DE AUTENTICACIÓN
# ==========================================

@app.post("/api/register", response_model=Usuario, status_code=201)
def register(user: UserCreate, db: Session = Depends(get_db)):
    # Verificar si existe
    db_user = db.query(UsuarioDB).filter(UsuarioDB.username == user.username).first()
    if db_user:
        raise HTTPException(status_code=400, detail="El usuario ya existe")
    
    hashed_pwd = seguridad.hashear_password(user.password)
    
    # ADAPTADO A TU DATABASE.PY: Usamos 'rol' en lugar de 'role'
    nuevo_usuario = UsuarioDB(
        username=user.username,
        hashed_password=hashed_pwd,
        rol=user.role,   # <--- AQUÍ ESTABA EL ERROR ANTES
        force_change_password=False
    )
    db.add(nuevo_usuario)
    db.commit()
    db.refresh(nuevo_usuario)
    return Usuario(username=user.username)

@app.post("/api/login")
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(UsuarioDB).filter(UsuarioDB.username == form_data.username).first()
    if not user:
        raise HTTPException(status_code=400, detail="Usuario o contraseña incorrectos")
    
    if not seguridad.verificar_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Usuario o contraseña incorrectos")
    
    # Generar token (En el token usamos 'role' estándar, pero lo sacamos de user.rol)
    access_token = seguridad.crear_token_acceso(data={"sub": user.username, "role": user.rol})
    
    return {
        "access_token": access_token, 
        "token_type": "bearer",
        "role": user.rol, # Devolvemos 'role' al frontend para que React entienda
        "username": user.username 
    }

# --- REEMPLAZA TODA LA FUNCIÓN cambiar_password CON ESTO ---

@app.post("/api/auth/change-password")
def cambiar_password(datos: PasswordChange, db: Session = Depends(get_db), u=Depends(get_current_user)):
    # 1. Obtenemos el nombre de usuario (sea dict u objeto)
    username = None
    if isinstance(u, dict):
        username = u.get("sub") or u.get("username")
    else:
        username = u.username

    # 2. Buscamos al usuario REAL en la Base de Datos
    user_db = db.query(UsuarioDB).filter(UsuarioDB.username == username).first()
    
    if not user_db:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # 3. Verificamos la contraseña usando el usuario de la DB (user_db), NO el token (u)
    if not seguridad.verificar_password(datos.current_password, user_db.hashed_password):
        raise HTTPException(status_code=400, detail="La contraseña actual es incorrecta")
    
    # 4. Actualizamos
    user_db.hashed_password = seguridad.hashear_password(datos.new_password)
    user_db.force_change_password = False 
    db.commit()
    
    return {"mensaje": "Contraseña actualizada correctamente"}
# ==========================================
# 5. ENDPOINTS DEL SISTEMA
# ==========================================

@app.post("/api/generar_horario")
def generar_horario_automatico(db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    # 1. Limpiar asignaciones anteriores
    db.query(AsignacionDB).delete()
    db.commit()
    
    # 2. Obtener configuraciones globales (ej: Almuerzos bloqueados)
    config_db = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "preferencias_horarios").first()
    almuerzo_slots = set()
    if config_db:
        try: almuerzo_slots = set(json.loads(config_db.value_json).get("almuerzo_slots", []))
        except: pass

    # 3. Obtener todos los REQUISITOS (Qué materias hay que dar)
    # Ordenamos: Primero los que tienen Aula Fija (más difícil), luego por cantidad de horas (de mayor a menor)
    requisitos = db.query(RequisitoDB).all()
    requisitos.sort(key=lambda r: (1 if r.aula_preferida_id else 0, r.horas_semanales), reverse=True)
    
    # --- CACHÉ DE DISPONIBILIDAD DE PROFESORES ---
    # Convertimos la lista JSON ["Lunes-07:40", ...] en un Set de Python para búsqueda rápida
    profesores_cache = {}
    todos_profes = db.query(ProfesorDB).all()
    for p in todos_profes:
        try:
            # Si tiene disponibilidad cargada, la usamos.
            # Si es NULL o vacía, asumimos que NO cargó nada (¿Lo tratamos como Libre o como No Disponible?)
            # En este caso: Si no cargó nada, asumimos que está LIBRE (dispo_list vacía = no restricciones)
            # PERO: Tu frontend manda una lista con lo que SÍ puede.
            # Ajuste: Si la lista tiene elementos, SOLO puede en esos elementos.
            # Si la lista está vacía, asumimos que puede siempre (o nunca, según tu política).
            # Vamos a asumir: Si tiene datos, respetamos. Si no, libre.
            
            dispo_list = json.loads(p.disponibilidad_json) if p.disponibilidad_json else []
            profesores_cache[p.id] = set(dispo_list)
        except:
            profesores_cache[p.id] = set()

    asignaciones_creadas = 0
    conflictos_log = []

    # 4. EL ALGORITMO PRINCIPAL
    for req in requisitos:
        horas_pendientes = req.horas_semanales
        
        # Intentamos ubicar cada hora
        for dia in DIAS_SEMANA:
            if horas_pendientes <= 0: break
            
            for hora in HORARIOS_ORDENADOS:
                if horas_pendientes <= 0: break
                
                clave_tiempo = f"{dia}-{hora}"
                
                # A. REGLA: NO ALMUERZOS
                if hora in almuerzo_slots: continue
                
                # B. REGLA: DISPONIBILIDAD DEL DOCENTE (NUEVO) 🧠
                if req.profesor_id:
                    dispo_del_profe = profesores_cache.get(req.profesor_id)
                    # Si el set tiene datos, significa que el profe marcó sus horas permitidas.
                    # Si la hora actual NO está en ese set, saltamos.
                    if dispo_del_profe and len(dispo_del_profe) > 0:
                        if clave_tiempo not in dispo_del_profe:
                            continue # El profe dijo que no.

                # C. REGLA: CURSO LIBRE
                ocupado_curso = db.query(AsignacionDB).filter(
                    AsignacionDB.curso_id == req.curso_id, 
                    AsignacionDB.dia == dia, 
                    AsignacionDB.hora_rango == hora
                ).first()
                if ocupado_curso: continue

                # D. REGLA: PROFESOR LIBRE (UBICUIDAD)
                if req.profesor_id:
                    ocupado_profe = db.query(AsignacionDB).filter(
                        AsignacionDB.profesor_id == req.profesor_id,
                        AsignacionDB.dia == dia,
                        AsignacionDB.hora_rango == hora
                    ).first()
                    if ocupado_profe: continue
                
                # E. REGLA: AULA
                aula_asignada_id = None
                
                # Si pide aula específica
                if req.aula_preferida_id:
                    ocupado_aula = db.query(AsignacionDB).filter(
                        AsignacionDB.aula_id == req.aula_preferida_id,
                        AsignacionDB.dia == dia,
                        AsignacionDB.hora_rango == hora
                    ).first()
                    if ocupado_aula: continue
                    aula_asignada_id = req.aula_preferida_id
                
                # --- SI PASÓ TODOS LOS FILTROS: ASIGNAMOS ---
                nueva_asignacion = AsignacionDB(
                    id=f"asig-{uuid.uuid4()}",
                    dia=dia, hora_rango=hora,
                    curso_id=req.curso_id, materia_id=req.materia_id,
                    profesor_id=req.profesor_id, aula_id=aula_asignada_id 
                )
                db.add(nueva_asignacion)
                # Guardamos para que la siguiente iteración sepa que esto ya se ocupó
                db.commit() 
                
                horas_pendientes -= 1
                asignaciones_creadas += 1

        if horas_pendientes > 0:
            conflictos_log.append(f"Materia {req.materia_id} (Curso {req.curso_id}): Faltaron asignar {horas_pendientes} hs.")

    mensaje_final = f"¡Proceso finalizado! 🚀\nSe generaron {asignaciones_creadas} módulos."
    if conflictos_log:
        mensaje_final += f"\n\n⚠️ Conflictos:\n" + "\n".join(conflictos_log)

    return {"mensaje": mensaje_final}

# --- REQUISITOS ---
@app.get("/api/requisitos")
def list_all_requisitos(db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    resultados = db.query(RequisitoDB).options(
        joinedload(RequisitoDB.curso), joinedload(RequisitoDB.materia),
        joinedload(RequisitoDB.profesor), joinedload(RequisitoDB.aula_preferida)
    ).all()
    lista_visual = []
    for r in resultados:
        lista_visual.append({
            "id": r.id,
            "curso_id": r.curso_id, "curso_anio": r.curso.anio if r.curso else "?", "curso_division": r.curso.division if r.curso else "?",
            "materia_id": r.materia_id, "materia_nombre": r.materia.nombre if r.materia else "??",
            "materia_color": r.materia.color_hex if r.materia else "#cccccc",
            "profesor_id": r.profesor_id, "profesor_nombre": r.profesor.nombre if r.profesor else "Sin Asignar",
            "aula_nombre": r.aula_preferida.nombre if r.aula_preferida else None,
            "horas_semanales": r.horas_semanales
        })
    return lista_visual

@app.post("/api/requisitos")
def add_req(r: RequisitoCreate, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    nuevo_req = RequisitoDB(
        id=f"req-{uuid.uuid4()}", curso_id=r.curso_id, materia_id=r.materia_id,
        profesor_id=r.profesor_id, aula_preferida_id=r.aula_preferida_id, horas_semanales=r.horas_semanales
    )
    db.add(nuevo_req)
    try: db.commit(); db.refresh(nuevo_req); return {"mensaje": "Creado", "id": nuevo_req.id}
    except Exception as e: db.rollback(); raise HTTPException(400, str(e))

@app.delete("/api/requisitos/{id}")
def delete_req(id: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    req = db.query(RequisitoDB).filter(RequisitoDB.id == id).first()
    if not req: raise HTTPException(404)
    db.delete(req); db.commit()
    return {"mensaje": "Eliminado"}

@app.get("/api/profesores", response_model=List[Profesor])
def obtener_profesores(db: Session = Depends(get_db), u=Depends(get_current_user)):
    # DEVOLVEMOS EL COLOR
    return [
        Profesor(
            id=p.id, 
            nombre=p.nombre, 
            disponibilidad=json.loads(p.disponibilidad_json or '[]'),
            color=p.color or "#0d9488" # <--- NUEVO
        ) for p in db.query(ProfesorDB).all()
    ]

# Busca el endpoint @app.post("/api/profesores") y actualízalo:
@app.post("/api/profesores", response_model=Profesor, status_code=201)
def agregar_profesor(p: Profesor, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    if db.query(ProfesorDB).filter(ProfesorDB.nombre == p.nombre).first(): raise HTTPException(409, "Existe")
    pid = f"p-{uuid.uuid4()}"
    
    db.add(ProfesorDB(
        id=pid, 
        nombre=p.nombre, 
        dni=p.dni, # <--- GUARDAMOS EL DNI
        disponibilidad_json=json.dumps(p.disponibilidad),
        color=p.color
    ))
    
    # Creamos usuario para login
    if not db.query(UsuarioDB).filter(UsuarioDB.username == p.nombre).first():
        db.add(UsuarioDB(username=p.nombre, hashed_password=seguridad.hashear_password("1234"), rol="profesor", force_change_password=True))
    
    db.commit()
    return p

@app.delete("/api/profesores/{pid}", status_code=204)
def borrar_profesor(pid: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    p = db.query(ProfesorDB).filter(ProfesorDB.id == pid).first()
    if not p: raise HTTPException(404)
    u_db = db.query(UsuarioDB).filter(UsuarioDB.username == p.nombre).first()
    if u_db: db.delete(u_db)
    db.delete(p); db.commit()
    return Response(status_code=204)

# --- MATERIAS ---
@app.get("/api/materias", response_model=List[Materia])
def get_materias(db: Session = Depends(get_db), u=Depends(get_current_user)):
    return [Materia(id=m.id, nombre=m.nombre, color_hex=m.color_hex) for m in db.query(MateriaDB).all()]

@app.post("/api/materias", response_model=Materia)
def add_materia(m: Materia, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    if db.query(MateriaDB).filter(MateriaDB.nombre == m.nombre).first(): raise HTTPException(409)
    nm = MateriaDB(id=f"m-{uuid.uuid4()}", nombre=m.nombre, color_hex=m.color_hex)
    db.add(nm); db.commit(); db.refresh(nm)
    return Materia(id=nm.id, nombre=nm.nombre, color_hex=nm.color_hex)

@app.delete("/api/materias/{mid}", status_code=204)
def del_materia(mid: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    db.query(MateriaDB).filter(MateriaDB.id == mid).delete(); db.commit()
    return Response(status_code=204)

# --- CURSOS ---
@app.get("/api/cursos", response_model=List[Curso])
def get_cursos(db: Session = Depends(get_db), u=Depends(get_current_user)):
    return [Curso(id=c.id, anio=c.anio, division=c.division, cantidad_alumnos=c.cantidad_alumnos, nombre_display=c.nombre_completo) for c in db.query(CursoDB).all()]

@app.post("/api/cursos", response_model=Curso)
def add_curso(c: Curso, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    if db.query(CursoDB).filter(CursoDB.anio == c.anio, CursoDB.division == c.division).first(): raise HTTPException(409)
    nc = CursoDB(id=f"c-{uuid.uuid4()}", anio=c.anio, division=c.division, cantidad_alumnos=c.cantidad_alumnos)
    db.add(nc); db.commit(); db.refresh(nc)
    return Curso(id=nc.id, anio=nc.anio, division=nc.division, cantidad_alumnos=nc.cantidad_alumnos, nombre_display=nc.nombre_completo)

@app.delete("/api/cursos/{cid}", status_code=204)
def del_curso(cid: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    db.query(CursoDB).filter(CursoDB.id == cid).delete(); db.commit()
    return Response(status_code=204)

# --- AULAS ---
@app.get("/api/aulas", response_model=List[Aula])
def get_aulas(db: Session = Depends(get_db), u=Depends(get_current_user)):
    return [Aula(id=a.id, nombre=a.nombre, tipo=a.tipo, capacidad=a.capacidad) for a in db.query(AulaDB).all()]

@app.post("/api/aulas", response_model=Aula)
def add_aula(a: Aula, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    if db.query(AulaDB).filter(AulaDB.nombre == a.nombre).first(): raise HTTPException(409)
    na = AulaDB(id=f"a-{uuid.uuid4()}", nombre=a.nombre, tipo=a.tipo, capacidad=a.capacidad)
    db.add(na); db.commit(); db.refresh(na)
    return Aula(id=na.id, nombre=na.nombre, tipo=na.tipo, capacidad=na.capacidad)

@app.delete("/api/aulas/{aid}", status_code=204)
def del_aula(aid: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    db.query(AulaDB).filter(AulaDB.id == aid).delete(); db.commit()
    return Response(status_code=204)

# --- HORARIOS & EXPORT ---
@app.get("/api/horarios/{cid}")
def get_horario_curso(cid: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    asigs = db.query(AsignacionDB).options(
        joinedload(AsignacionDB.profesor), joinedload(AsignacionDB.materia), joinedload(AsignacionDB.aula)
    ).filter(AsignacionDB.curso_id == cid).all()
    vista = {}
    for a in asigs:
        vista.setdefault(a.hora_rango, {})[a.dia] = {
            "text": f"{a.materia.nombre if a.materia else '??'}",
            "profesor_nombre": a.profesor.nombre if a.profesor else "Sin Profe",
            "materia_nombre": a.materia.nombre if a.materia else "??",
            "color_materia": a.materia.color_hex if a.materia else "#0d9488",
            "id": a.id, "aula_nombre": a.aula.nombre if a.aula else "Sin Aula"
        }
    return vista

@app.get("/api/export/excel")
def export_excel(db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    wb = openpyxl.Workbook(); 
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    cursos = db.query(CursoDB).all()
    asigs = db.query(AsignacionDB).options(joinedload(AsignacionDB.profesor), joinedload(AsignacionDB.materia), joinedload(AsignacionDB.aula)).all()
    
    for c in cursos:
        ws = wb.create_sheet(f"{c.nombre_completo}"[:30])
        ws.append(['Hora'] + DIAS_SEMANA)
        for cell in ws[1]: 
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1D72B8")
        
        vista = {}
        for a in asigs:
            if a.curso_id == c.id:
                prof = a.profesor.nombre if a.profesor else "??"
                mat = a.materia.nombre if a.materia else "??"
                aula = f" [{a.aula.nombre}]" if a.aula else ""
                vista.setdefault(a.hora_rango, {})[a.dia] = f"{mat}\n({prof}){aula}"
        
        for h in HORARIOS_ORDENADOS:
            row = [h] + [vista.get(h, {}).get(d, "") for d in DIAS_SEMANA]
            ws.append(row)
            
        ws.column_dimensions['A'].width = 15
        for col in ['B','C','D','E','F']: 
            ws.column_dimensions[col].width = 25
            
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.getvalue(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=Horarios.xlsx"})


# --- CONFIG Y HERRAMIENTAS ---
@app.get("/api/config/preferencias")
def obtener_preferencias(db: Session = Depends(get_db)):
    conf = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "preferencias_horarios").first()
    if not conf: return {"almuerzo_slots": []}
    try: return json.loads(conf.value_json)
    except: return {"almuerzo_slots": []}

@app.post("/api/config/preferencias")
def guardar_preferencias(prefs: Preferencias, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    conf = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "preferencias_horarios").first()
    if not conf:
        conf = ConfiguracionDB(key="preferencias_horarios")
        db.add(conf)
    conf.value_json = json.dumps({"almuerzo_slots": prefs.almuerzo_slots})
    db.commit()
    return {"mensaje": "Guardado"}

@app.get("/api/reportes/carga-horaria-profesor", response_model=List[ReporteCargaHoraria])
def reporte_carga(db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    # Consultamos Nombre, Color y Cantidad
    res = db.query(
        ProfesorDB.nombre, 
        ProfesorDB.color, 
        func.count(AsignacionDB.id).label("cnt")
    ).join(AsignacionDB, ProfesorDB.id == AsignacionDB.profesor_id, isouter=True)\
     .group_by(ProfesorDB.id)\
     .order_by(func.count(AsignacionDB.id).desc()).all()
    
    return [
        ReporteCargaHoraria(
            nombre_profesor=n, 
            horas_asignadas=c, 
            color=col or "#0d9488" # <--- Usamos el color del profe
        ) for n, col, c in res
    ]

@app.delete("/api/admin/reset-horarios", status_code=200)
def reset_assignments(db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    db.query(AsignacionDB).delete(); db.commit()
    return {"mensaje": "Horarios eliminados."}

@app.post("/api/admin/reset-password/{username}")
def reset_password_admin(username: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    user = db.query(UsuarioDB).filter(UsuarioDB.username == username).first()
    if not user: raise HTTPException(404, "Usuario no encontrado")
    user.hashed_password = seguridad.hashear_password("1234")
    user.force_change_password = True
    db.commit()
    return {"mensaje": f"Clave de '{username}' restablecida a '1234'"}

@app.post("/api/profesores/buscar-suplentes")
def buscar_suplentes(req: BusquedaSuplente, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    hora_rango = calcular_hora_rango(req.hora_inicio)
    if not hora_rango: raise HTTPException(400, "Hora inválida")
    slot_buscado = f"{req.dia}-{req.hora_inicio}"
    profesores = db.query(ProfesorDB).all()
    ocupados = db.query(AsignacionDB.profesor_id).filter(AsignacionDB.dia == req.dia, AsignacionDB.hora_rango == hora_rango).all()
    ids_ocupados = {ocup.profesor_id for ocup in ocupados}
    disponibles = []
    for p in profesores:
        try: disp_set = set(json.loads(p.disponibilidad_json or '[]'))
        except: disp_set = set()
        if slot_buscado in disp_set and p.id not in ids_ocupados:
            disponibles.append({"id": p.id, "nombre": p.nombre})
    return disponibles

@app.post("/api/horarios/mover")
def mover_asignacion(mov: MovimientoHorario, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    # 1. Buscamos la asignación que se quiere mover ("El Viajero")
    viajero = db.query(AsignacionDB).filter(AsignacionDB.id == mov.asignacion_id).first()
    if not viajero:
        raise HTTPException(404, "Asignación no encontrada")

    # Datos del destino
    nuevo_dia = mov.nuevo_dia
    nueva_hora = mov.nueva_hora
    curso_actual_id = viajero.curso_id

    # --- 🛡️ REGLA 1: VALIDAR DISPONIBILIDAD DEL DOCENTE ---
    if viajero.profesor_id:
        profesor = db.query(ProfesorDB).filter(ProfesorDB.id == viajero.profesor_id).first()
        if profesor and profesor.disponibilidad_json:
            try:
                dispo = json.loads(profesor.disponibilidad_json)
                slot_destino = f"{nuevo_dia}-{nueva_hora}"
                # Si la lista de disponibilidad NO está vacía y el slot NO está en ella...
                if dispo and (slot_destino not in dispo):
                    raise HTTPException(
                        status_code=409, 
                        detail=f"El profesor {profesor.nombre} NO tiene disponibilidad el {nuevo_dia} a las {nueva_hora}."
                    )
            except json.JSONDecodeError:
                pass # Si falla el JSON, ignoramos esta regla por seguridad

    # --- 🛡️ REGLA 2: VALIDAR QUE EL DOCENTE NO ESTÉ EN OTRO CURSO ---
    if viajero.profesor_id:
        # Buscamos si existe ALGUNA asignación de este profe, en ese día y hora, PERO en otro ID (para no contarse a sí mismo)
        ocupado = db.query(AsignacionDB).options(joinedload(AsignacionDB.curso)).filter(
            AsignacionDB.profesor_id == viajero.profesor_id,
            AsignacionDB.dia == nuevo_dia,
            AsignacionDB.hora_rango == nueva_hora,
            AsignacionDB.id != viajero.id # Importante: que no sea él mismo
        ).first()

        if ocupado:
            # Si encontramos algo, es un choque.
            nombre_curso_choque = f"{ocupado.curso.anio} {ocupado.curso.division}" if ocupado.curso else "otro curso"
            raise HTTPException(
                status_code=409, 
                detail=f"CONFLICTO: El profesor ya está asignado en {nombre_curso_choque} a esa hora."
            )

    # --- LÓGICA DE MOVIMIENTO (Si pasó las validaciones) ---
    
    # Verificamos quién está en el destino (el "Inquilino") para hacer enroque
    inquilino = db.query(AsignacionDB).filter(
        AsignacionDB.curso_id == viajero.curso_id,
        AsignacionDB.dia == nuevo_dia,
        AsignacionDB.hora_rango == nueva_hora
    ).first()

    mensaje = "Horario movido exitosamente"
    
    # Guardamos origen por si hay swap
    dia_origen = viajero.dia
    hora_origen = viajero.hora_rango

    if inquilino:
        # CASO SWAP: Hay alguien. Lo mandamos al origen del viajero.
        # ¡OJO! Deberíamos validar si el "Inquilino" puede ir al origen, pero por ahora lo permitimos para facilitar el uso.
        inquilino.dia = dia_origen
        inquilino.hora_rango = hora_origen
        mensaje = "Horarios intercambiados (Swap)"

    # Movemos al viajero
    viajero.dia = nuevo_dia
    viajero.hora_rango = nueva_hora
    
    db.commit()
    return {"mensaje": mensaje}

@app.get("/api/horarios/profesor/me")
def obtener_mis_horarios(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    nombre_profe = current_user["username"]
    profe_db = db.query(ProfesorDB).filter(ProfesorDB.nombre == nombre_profe).first()
    
    if not profe_db:
        # En tu DB vieja no hay 'profesor_nombre', pero podemos buscar por relación si existiera.
        # En tu caso, los profesores tienen el mismo nombre de usuario que el nombre en ProfesorDB.
        # Si no se encuentra el profesor en DB, no hay mucho que hacer, devolvemos vacío.
        return {}
    
    asignaciones = db.query(AsignacionDB).filter(AsignacionDB.profesor_id == profe_db.id).all()
    
    grilla = {}
    for a in asignaciones:
        if a.hora_rango not in grilla: grilla[a.hora_rango] = {}
        grilla[a.hora_rango][a.dia] = {
            "materia": a.materia.nombre if a.materia else "??",
            "curso": f"{a.curso.anio} \"{a.curso.division}\"" if a.curso else "?",
            "aula": a.aula.nombre if a.aula else "Sin Aula",
            "color": a.materia.color_hex if a.materia else "#0d9488"
        }
    return grilla

# Modelo para recibir la configuración
class ConfiguracionInstitucional(BaseModel):
    nombre: str
    direccion: Optional[str] = ""
    logo_base64: Optional[str] = None 

# ==========================================
# 6. CONFIGURACIÓN INSTITUCIONAL (DEBUG MODE)
# ==========================================

# Modelo para recibir la configuración
class ConfiguracionInstitucional(BaseModel):
    nombre: str
    direccion: Optional[str] = ""
    logo_base64: Optional[str] = None 

@app.post("/api/config/institucion")
def guardar_config_institucion(config: ConfiguracionInstitucional, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    print("--- 🟢 INICIANDO GUARDADO DE CONFIGURACIÓN ---")
    print(f"Recibido Nombre: {config.nombre}")
    print(f"Recibido Dirección: {config.direccion}")
    if config.logo_base64:
        print(f"Recibido Logo (Tamaño): {len(config.logo_base64)} caracteres")
    else:
        print("Recibido Logo: None")

    try:
        # 1. Guardar Nombre
        conf_nombre = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "institucion_nombre").first()
        if not conf_nombre:
            print("-> Creando registro 'institucion_nombre'...")
            conf_nombre = ConfiguracionDB(key="institucion_nombre")
            db.add(conf_nombre)
        conf_nombre.value_json = json.dumps(config.nombre)

        # 2. Guardar Direccion
        conf_dir = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "institucion_direccion").first()
        if not conf_dir:
            print("-> Creando registro 'institucion_direccion'...")
            conf_dir = ConfiguracionDB(key="institucion_direccion")
            db.add(conf_dir)
        conf_dir.value_json = json.dumps(config.direccion)

        # 3. Guardar Logo
        if config.logo_base64:
            print("-> Procesando guardado de Logo...")
            conf_logo = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "institucion_logo").first()
            if not conf_logo:
                print("-> Creando registro 'institucion_logo'...")
                conf_logo = ConfiguracionDB(key="institucion_logo")
                db.add(conf_logo)
            conf_logo.value_json = json.dumps(config.logo_base64)
            print("-> Logo asignado a la base de datos.")

        print("-> Ejecutando COMMIT...")
        db.commit()
        print("--- ✅ GUARDADO EXITOSO ---")
        return {"mensaje": "Configuración institucional guardada"}

    except Exception as e:
        print(f"--- ❌ ERROR GRAVE AL GUARDAR: {e} ---")
        db.rollback() # Deshacer cambios si falla
        raise HTTPException(status_code=500, detail=f"Error interno: {str(e)}")

@app.get("/api/config/institucion")
def obtener_config_institucion(db: Session = Depends(get_db)):
    try:
        c_nombre = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "institucion_nombre").first()
        c_dir = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "institucion_direccion").first()
        c_logo = db.query(ConfiguracionDB).filter(ConfiguracionDB.key == "institucion_logo").first()
        
        # Logs para ver qué estamos recuperando
        # print(f"GET Nombre encontrado: {c_nombre.value_json if c_nombre else 'No'}")

        return {
            "nombre": json.loads(c_nombre.value_json) if c_nombre else "",
            "direccion": json.loads(c_dir.value_json) if c_dir else "",
            "logo_base64": json.loads(c_logo.value_json) if c_logo else None
        }
    except Exception as e:
        print(f"Error al leer configuración: {e}")
        return {"nombre": "", "direccion": "", "logo_base64": None}
    
    # --- AGREGAR EN main.py (Junto a los otros endpoints de horarios) ---

@app.get("/api/horarios/profesor/{pid}")
def get_horario_profesor_admin(pid: str, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    # Buscamos todas las asignaciones de ESTE profesor
    asigs = db.query(AsignacionDB).options(
        joinedload(AsignacionDB.curso), 
        joinedload(AsignacionDB.materia), 
        joinedload(AsignacionDB.aula)
    ).filter(AsignacionDB.profesor_id == pid).all()
    
    grilla = {}
    for a in asigs:
        # Estructura: grilla[hora][dia] = "Materia (Curso) [Aula]"
        if a.hora_rango not in grilla: grilla[a.hora_rango] = {}
        
        texto_curso = f"{a.curso.anio} {a.curso.division}" if a.curso else "?"
        texto_materia = a.materia.nombre if a.materia else "?"
        texto_aula = f" [{a.aula.nombre}]" if a.aula else ""
        
        grilla[a.hora_rango][a.dia] = {
            "texto": f"{texto_materia}\n({texto_curso}){texto_aula}",
            "materia": texto_materia,
            "curso": texto_curso,
            "aula": a.aula.nombre if a.aula else "-"
        }
    return grilla

# --- AGREGAR EN main.py (Sección Profesores) ---

@app.put("/api/profesores/{pid}")
def actualizar_profesor(pid: str, p: Profesor, db: Session = Depends(get_db), u=Depends(get_current_admin_user)):
    profe = db.query(ProfesorDB).filter(ProfesorDB.id == pid).first()
    if not profe: 
        raise HTTPException(status_code=404, detail="Profesor no encontrado")
    
    # Actualizamos los datos
    profe.nombre = p.nombre
    profe.dni = p.dni
    profe.color = p.color
    profe.disponibilidad_json = json.dumps(p.disponibilidad)
    
    db.commit()
    return {"mensaje": "Profesor actualizado correctamente"}

# --- AGREGAR EN main.py (Sección Auth) ---

class PasswordChange(BaseModel):
    old_password: str
    new_password: str
